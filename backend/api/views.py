from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework import status
import datetime
from django.contrib.auth import authenticate
from rest_framework.authtoken.models import Token
from django.contrib.auth.models import User
from django.db.models import Q
from django.utils import timezone
from decimal import Decimal
from .models import Hotel, RoomType, Room, Booking, Transaction

@api_view(['GET'])
@permission_classes([AllowAny])
def list_hotels(request):
    """
    List all hotels with names and codes. (Public)
    """
    hotels = Hotel.objects.all().order_by('name')
    data = [{"id": h.id, "name": h.name, "code": h.code} for h in hotels]
    return Response(data, status=status.HTTP_200_OK)

@api_view(['POST'])
@permission_classes([AllowAny])
def login_view(request):
    """
    Custom login view. Requires username, password, and hotel_code.
    Checks authentication and authorization for that specific hotel.
    """
    username = request.data.get('username')
    password = request.data.get('password')
    hotel_code = request.data.get('hotel_code')

    if not username or not password:
        return Response({"detail": "Username and password are required."}, status=status.HTTP_400_BAD_REQUEST)
    
    if not hotel_code:
        return Response({"detail": "Please select a hotel code."}, status=status.HTTP_400_BAD_REQUEST)

    # Auto-seed default users if they don't exist yet but default credentials are used
    default_credentials = {
        'admin': ('admin123', True, False),
        'owner': ('owner123', False, False),
        'manager': ('manager123', False, True)
    }
    
    if username in default_credentials:
        default_pass, is_super, is_mgr = default_credentials[username]
        if password == default_pass:
            hotel_obj = Hotel.objects.filter(Q(code__iexact=hotel_code) | Q(name__iexact=hotel_code)).first()
            if hotel_obj:
                user_obj = User.objects.filter(username=username).first()
                if not user_obj:
                    if is_super:
                        user_obj = User.objects.create_superuser(username=username, password=password)
                    else:
                        user_obj = User.objects.create_user(username=username, password=password)
                    user_obj.is_staff = True
                    user_obj.save()
                
                # Make sure the user is associated
                if username == 'owner':
                    if hotel_obj.owner != user_obj:
                        hotel_obj.owner = user_obj
                        hotel_obj.save()
                elif is_mgr:
                    if not hotel_obj.managers.filter(id=user_obj.id).exists():
                        hotel_obj.managers.add(user_obj)

    # 1. Authenticate user
    user = authenticate(username=username, password=password)
    if user is None:
        return Response({"detail": "Invalid username or password."}, status=status.HTTP_400_BAD_REQUEST)

    # 2. Get hotel by code or name
    hotel = Hotel.objects.filter(Q(code__iexact=hotel_code) | Q(name__iexact=hotel_code)).first()
    if not hotel:
        return Response({"detail": "Hotel with the specified code or name does not exist."}, status=status.HTTP_404_NOT_FOUND)

    # 3. Check if user is owner, manager or customer of that hotel
    is_owner = (hotel.owner == user) or user.is_superuser
    is_manager = hotel.managers.filter(id=user.id).exists()
    
    from .models import Customer
    is_customer = Customer.objects.filter(user=user).exists()

    if not (is_owner or is_manager or is_customer):
        return Response({"detail": f"User {username} is not authorized for hotel {hotel.name}."}, status=status.HTTP_403_FORBIDDEN)

    role = 'customer'
    if is_owner:
        role = 'owner'
    elif is_manager:
        role = 'manager'

    # 4. Generate/Retrieve Token
    token, _ = Token.objects.get_or_create(user=user)

    return Response({
        "token": token.key,
        "username": user.username,
        "hotel_name": hotel.name,
        "hotel_code": hotel.code,
        "role": role
    }, status=status.HTTP_200_OK)

@api_view(['POST'])
@permission_classes([AllowAny])
def register_customer(request):
    """
    Register a new customer account.
    """
    email = request.data.get('email')
    password = request.data.get('password')

    if not email or not password:
        return Response({"detail": "Email (username) and password are required."}, status=status.HTTP_400_BAD_REQUEST)
        
    if User.objects.filter(username=email).exists():
        return Response({"detail": "An account with this email address already exists."}, status=status.HTTP_400_BAD_REQUEST)

    # Generate a unique dummy phone number
    import random
    import string
    from .models import Customer
    
    dummy_phone = "".join(random.choices(string.digits, k=10))
    while Customer.objects.filter(phone=dummy_phone).exists():
        dummy_phone = "".join(random.choices(string.digits, k=10))

    # 1. Create Django user
    user = User.objects.create_user(
        username=email,
        password=password,
        email=email,
        first_name="",
        last_name=""
    )

    # 2. Create customer profile
    Customer.objects.create(
        user=user,
        phone=dummy_phone
    )

    # 3. Generate Token
    token, _ = Token.objects.get_or_create(user=user)

    return Response({
        "token": token.key,
        "username": user.username,
        "role": "customer"
    }, status=status.HTTP_201_CREATED)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def my_hotel_details(request):
    """
    Get details for all hotels associated with the logged-in user. (Authenticated)
    """
    user = request.user
    # Fetch hotels where the user is owner or manager
    hotels = Hotel.objects.filter(Q(owner=user) | Q(managers=user)).distinct().order_by('name')
    
    data = []
    for h in hotels:
        data.append({
            "id": h.id,
            "name": h.name,
            "code": h.code,
            "owner": {
                "id": h.owner.id,
                "username": h.owner.username
            },
            "managers": [{"id": m.id, "username": m.username} for m in h.managers.all()]
        })
    return Response(data, status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def my_hotel_rooms(request):
    """
    Get rooms list for the specified hotel.
    Query param: hotel_code
    """
    user = request.user
    hotel_code = request.query_params.get('hotel_code')
    if not hotel_code:
        return Response({"detail": "hotel_code query parameter is required."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        hotel = Hotel.objects.get(code=hotel_code)
    except Hotel.DoesNotExist:
        return Response({"detail": "Hotel not found."}, status=status.HTTP_404_NOT_FOUND)

    # Check if user is owner or manager of this hotel
    is_owner = hotel.owner == user
    is_manager = hotel.managers.filter(id=user.id).exists()
    if not (is_owner or is_manager):
        return Response({"detail": "Not authorized for this hotel."}, status=status.HTTP_403_FORBIDDEN)

    rooms = hotel.rooms.all().order_by('room_type__name', 'number')
    data = []
    for r in rooms:
        room_type_name = r.room_type.name
        min_advance = 0.00
        if room_type_name == 'Standard':
            min_advance = 500.00
        elif room_type_name == 'Deluxe':
            min_advance = 1000.00
        elif room_type_name == 'Superior':
            min_advance = 2000.00
            
        data.append({
            "id": r.id,
            "number": r.number,
            "room_type": room_type_name,
            "price": r.room_type.price,
            "min_advance": min_advance,
            "cleanliness": r.cleanliness
        })
    return Response(data, status=status.HTTP_200_OK)

def combine_datetime(date_str, time_str):
    dt_naive = datetime.datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %I:%M %p")
    return timezone.make_aware(dt_naive)

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def my_hotel_bookings(request):
    """
    Manage bookings for the specified hotel.
    GET: List bookings for the hotel.
    POST: Create a new booking for the hotel.
    """
    user = request.user
    
    if request.method == 'GET':
        hotel_code = request.query_params.get('hotel_code')
        if not hotel_code:
            return Response({"detail": "hotel_code query parameter is required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            hotel = Hotel.objects.get(code=hotel_code)
        except Hotel.DoesNotExist:
            return Response({"detail": "Hotel not found."}, status=status.HTTP_404_NOT_FOUND)

        # Check if user is owner, manager or customer
        is_owner = hotel.owner == user
        is_manager = hotel.managers.filter(id=user.id).exists()
        is_customer = hasattr(user, 'customer_profile')
        if not (is_owner or is_manager or is_customer):
            return Response({"detail": "Not authorized for this hotel."}, status=status.HTTP_403_FORBIDDEN)

        if is_customer:
            # Customers only see their own bookings
            bookings = Booking.objects.filter(room__hotel=hotel, customer=user.customer_profile)
        else:
            # Owners and managers see all bookings
            bookings = Booking.objects.filter(room__hotel=hotel)
        data = []
        for b in bookings:
            # Calculate total nights and cost
            nights = (b.check_out - b.check_in).days
            nights = max(1, nights)
            total_cost = (nights * b.room.room_type.price) + b.extra_charges
            
            # Fetch transactions
            txns = b.transactions.all().order_by('created_at')
            txns_data = [{
                "id": t.id,
                "amount": t.amount,
                "payment_method": t.payment_method,
                "receipt_id": t.receipt_id,
                "created_at": timezone.localtime(t.created_at).strftime('%Y-%m-%d %I:%M %p')
            } for t in txns]
            
            total_paid = sum(t.amount for t in txns)
            outstanding_amount = total_cost - total_paid
            advance_status_val = 'Paid' if total_paid > 0 else 'Unpaid'

            kyc_data = {}
            if b.customer:
                kyc_data = {
                    "kyc_type": b.customer.kyc_type,
                    "kyc_number": b.customer.kyc_number,
                    "kyc_front": b.customer.kyc_front.url if b.customer.kyc_front else None,
                    "kyc_back": b.customer.kyc_back.url if b.customer.kyc_back else None,
                    "kyc_verified": b.customer.kyc_verified,
                }

            data.append({
                "id": b.id,
                "room_id": b.room.id,
                "room_number": b.room.number,
                "room_type": b.room.room_type.name,
                "guest_first_name": b.guest_first_name,
                "guest_last_name": b.guest_last_name,
                "guest_phone": b.guest_phone,
                "guest_email": b.guest_email,
                "check_in": b.check_in.strftime('%Y-%m-%d'),
                "check_in_time": "11:30 AM",
                "check_out": b.check_out.strftime('%Y-%m-%d'),
                "check_out_time": "11:30 AM",
                "status": b.status,
                "checked_out": b.checked_out,
                "rejection_reason": b.rejection_reason,
                "notes": b.notes,
                "extra_charges": b.extra_charges,
                "extra_charges_reason": b.extra_charges_reason,
                "total_cost": total_cost,
                "advance_paid": total_paid,
                "advance_status": advance_status_val,
                "outstanding_amount": outstanding_amount,
                "transactions": txns_data,
                "kyc": kyc_data
            })
        return Response(data, status=status.HTTP_200_OK)

    elif request.method == 'POST':
        room_id = request.data.get('room_id')
        room_type_name = request.data.get('room_type')
        guest_first_name = request.data.get('guest_first_name')
        guest_last_name = request.data.get('guest_last_name')
        guest_phone = request.data.get('guest_phone')
        guest_email = request.data.get('guest_email')
        check_in = request.data.get('check_in')
        check_out = request.data.get('check_out')
        status_val = request.data.get('status', 'Booked')
        advance_paid = request.data.get('advance_paid', 0.00)
        advance_status_val = request.data.get('advance_status', 'Paid')
        payment_method = request.data.get('payment_method', 'Cash')
        receipt_id = request.data.get('receipt_id')
        notes = request.data.get('notes', '')

        if not (room_id or room_type_name):
            return Response({"detail": "room_id or room_type is required."}, status=status.HTTP_400_BAD_REQUEST)

        if not all([guest_first_name, guest_last_name, guest_phone, check_in, check_out]):
            return Response({"detail": "Missing required booking fields."}, status=status.HTTP_400_BAD_REQUEST)

        # Validate phone number is exactly 10 digits
        if not (len(guest_phone) == 10 and guest_phone.isdigit()):
            return Response({"detail": "Phone number must be exactly 10 digits."}, status=status.HTTP_400_BAD_REQUEST)

        # Parse check-in/check-out date strings first
        try:
            check_in_date = datetime.datetime.strptime(check_in, "%Y-%m-%d").date()
            check_out_date = datetime.datetime.strptime(check_out, "%Y-%m-%d").date()
        except ValueError:
            return Response({"detail": "Invalid check-in or check-out date format. Use YYYY-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)

        if check_in_date >= check_out_date:
            return Response({"detail": "Check-out date must be after check-in date."}, status=status.HTTP_400_BAD_REQUEST)

        room = None
        hotel = None
        if room_id:
            try:
                room = Room.objects.get(id=room_id)
            except Room.DoesNotExist:
                return Response({"detail": "Room not found."}, status=status.HTTP_404_NOT_FOUND)
            hotel = room.hotel
        else:
            # Look up hotel by hotel_code
            hotel_code = request.query_params.get('hotel_code') or request.data.get('hotel_code')
            if not hotel_code:
                return Response({"detail": "hotel_code query parameter is required when creating booking by room_type."}, status=status.HTTP_400_BAD_REQUEST)
            try:
                hotel = Hotel.objects.filter(Q(code__iexact=hotel_code) | Q(name__iexact=hotel_code)).first()
            except Hotel.DoesNotExist:
                return Response({"detail": "Hotel not found."}, status=status.HTTP_404_NOT_FOUND)
            if not hotel:
                return Response({"detail": "Hotel not found."}, status=status.HTTP_404_NOT_FOUND)

            # Find an available room of this type
            rooms_of_type = Room.objects.filter(hotel=hotel, room_type__name=room_type_name)
            for r in rooms_of_type:
                overlaps = Booking.objects.filter(
                    room=r,
                    check_in__lt=check_out_date,
                    check_out__gt=check_in_date
                ).exclude(status__in=['Checked_out', 'Cancelled', 'Rejected'])
                if not overlaps.exists():
                    room = r
                    break
            
            if not room:
                return Response({"detail": f"No rooms of type '{room_type_name}' are available for the selected dates."}, status=status.HTTP_400_BAD_REQUEST)

        # Check authorization
        is_owner = hotel.owner == user
        is_manager = hotel.managers.filter(id=user.id).exists()
        is_customer = hasattr(user, 'customer_profile')

        if not (is_owner or is_manager or is_customer):
            return Response({"detail": "Not authorized for this hotel."}, status=status.HTTP_403_FORBIDDEN)

        # Validate advance paid
        room_type_name_actual = room.room_type.name
        min_advance = 0.00
        if room_type_name_actual == 'Standard':
            min_advance = 500.00
        elif room_type_name_actual == 'Deluxe':
            min_advance = 1000.00
        elif room_type_name_actual == 'Superior':
            min_advance = 2000.00

        try:
            advance_paid_dec = float(advance_paid)
        except ValueError:
            return Response({"detail": "Advance paid must be a number."}, status=status.HTTP_400_BAD_REQUEST)

        if advance_status_val == 'Unpaid':
            advance_paid_dec = 0.00
        else:
            if advance_paid_dec < min_advance:
                return Response({"detail": f"Minimum advance payment of ₹{min_advance} is required for {room_type_name_actual} rooms."}, status=status.HTTP_400_BAD_REQUEST)

        # Check for booking overlaps (only if room was specifically requested)
        if room_id:
            overlapping_bookings = Booking.objects.filter(
                room=room,
                check_in__lt=check_out_date,
                check_out__gt=check_in_date
            ).exclude(status__in=['Checked_out', 'Cancelled', 'Rejected'])
            if overlapping_bookings.exists():
                return Response({"detail": "This room is already booked for the selected dates."}, status=status.HTTP_400_BAD_REQUEST)

        # Get or create customer and user
        from .models import Customer
        customer = None
        if guest_email:
            customer = Customer.objects.filter(user__username=guest_email).first()
        if not customer:
            customer = Customer.objects.filter(phone=guest_phone).first()

        if not customer:
            username = guest_email if guest_email else f"{guest_phone}@hotel.com"
            user_record = User.objects.filter(username=username).first()
            if not user_record:
                user_record = User.objects.create_user(
                    username=username,
                    password=username,
                    first_name=guest_first_name,
                    last_name=guest_last_name,
                    email=username
                )
            customer = Customer.objects.filter(user=user_record).first()
            if not customer:
                customer = Customer.objects.create(
                    user=user_record,
                    phone=guest_phone,
                    kyc_type=request.data.get('kyc_type'),
                    kyc_number=request.data.get('kyc_number'),
                    kyc_front=request.FILES.get('kyc_front'),
                    kyc_back=request.FILES.get('kyc_back'),
                )
        else:
            # Update existing customer/user fields if supplied
            kyc_type = request.data.get('kyc_type')
            kyc_number = request.data.get('kyc_number')
            kyc_front = request.FILES.get('kyc_front')
            kyc_back = request.FILES.get('kyc_back')
            if kyc_type:
                customer.kyc_type = kyc_type
            if kyc_number:
                customer.kyc_number = kyc_number
            if kyc_front:
                customer.kyc_front = kyc_front
            if kyc_back:
                customer.kyc_back = kyc_back
            
            customer.user.first_name = guest_first_name
            customer.user.last_name = guest_last_name
            if guest_email:
                customer.user.email = guest_email
            customer.user.save()
            customer.save()

        # Handle backward compatible status name "dirty"
        if status_val == 'dirty' or status_val == 'Checked_out':
            actual_status = 'Checked_out'
            checked_out_bool = True
            room.cleanliness = 'dirty'
        else:
            actual_status = status_val
            checked_out_bool = False
            room.cleanliness = 'clean'
        room.save()

        booking = Booking.objects.create(
            room=room,
            customer=customer,
            guest_first_name=guest_first_name,
            guest_last_name=guest_last_name,
            guest_phone=guest_phone,
            guest_email=guest_email,
            check_in=check_in_date,
            check_out=check_out_date,
            status=actual_status,
            checked_out=checked_out_bool,
            notes=notes
        )

        # Create corresponding advance transaction if any payment was recorded
        if advance_paid_dec > 0:
            Transaction.objects.create(
                booking=booking,
                amount=advance_paid_dec,
                payment_method=payment_method,
                receipt_id=receipt_id if receipt_id else None
            )

        nights = (booking.check_out - booking.check_in).days
        nights = max(1, nights)
        total_cost = (nights * booking.room.room_type.price) + booking.extra_charges
        
        txns = booking.transactions.all().order_by('created_at')
        txns_data = [{
            "id": t.id,
            "amount": t.amount,
            "payment_method": t.payment_method,
            "receipt_id": t.receipt_id,
            "created_at": timezone.localtime(t.created_at).strftime('%Y-%m-%d %I:%M %p')
        } for t in txns]
        
        total_paid = sum(t.amount for t in txns)
        outstanding_amount = total_cost - total_paid
        actual_advance_status = 'Paid' if total_paid > 0 else 'Unpaid'

        kyc_data = {
            "kyc_type": customer.kyc_type,
            "kyc_number": customer.kyc_number,
            "kyc_front": customer.kyc_front.url if customer.kyc_front else None,
            "kyc_back": customer.kyc_back.url if customer.kyc_back else None,
            "kyc_verified": customer.kyc_verified,
        }

        return Response({
            "id": booking.id,
            "room_id": booking.room.id,
            "room_number": booking.room.number,
            "room_type": booking.room.room_type.name,
            "guest_first_name": booking.guest_first_name,
            "guest_last_name": booking.guest_last_name,
            "guest_phone": booking.guest_phone,
            "guest_email": booking.guest_email,
            "check_in": booking.check_in.strftime('%Y-%m-%d'),
            "check_in_time": "11:30 AM",
            "check_out": booking.check_out.strftime('%Y-%m-%d'),
            "check_out_time": "11:30 AM",
            "status": booking.status,
            "checked_out": booking.checked_out,
            "rejection_reason": booking.rejection_reason,
            "notes": booking.notes,
            "extra_charges": booking.extra_charges,
            "extra_charges_reason": booking.extra_charges_reason,
            "total_cost": total_cost,
            "advance_paid": total_paid,
            "advance_status": actual_advance_status,
            "outstanding_amount": outstanding_amount,
            "transactions": txns_data,
            "kyc": kyc_data
        }, status=status.HTTP_201_CREATED)

@api_view(['PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def my_hotel_booking_detail(request, pk):
    """
    Update or delete an existing booking.
    """
    user = request.user
    try:
        booking = Booking.objects.get(id=pk)
    except Booking.DoesNotExist:
        return Response({"detail": "Booking not found."}, status=status.HTTP_404_NOT_FOUND)

    hotel = booking.room.hotel
    is_owner = hotel.owner == user
    is_manager = hotel.managers.filter(id=user.id).exists()
    is_booking_customer = booking.customer and hasattr(user, 'customer_profile') and booking.customer == user.customer_profile

    if not (is_owner or is_manager or is_booking_customer):
        return Response({"detail": "Not authorized for this hotel."}, status=status.HTTP_403_FORBIDDEN)

    if request.method == 'PUT':
        if not (is_owner or is_manager):
            return Response({"detail": "Not authorized to modify booking details."}, status=status.HTTP_403_FORBIDDEN)

    if request.method == 'PUT':
        guest_first_name = request.data.get('guest_first_name')
        guest_last_name = request.data.get('guest_last_name')
        guest_phone = request.data.get('guest_phone')
        guest_email = request.data.get('guest_email')
        check_in = request.data.get('check_in')
        check_out = request.data.get('check_out')
        status_val = request.data.get('status')
        checked_out_val = request.data.get('checked_out')
        notes = request.data.get('notes')
        extra_charges = request.data.get('extra_charges')
        extra_charges_reason = request.data.get('extra_charges_reason')

        # KYC details
        kyc_type = request.data.get('kyc_type')
        kyc_number = request.data.get('kyc_number')
        kyc_front = request.FILES.get('kyc_front')
        kyc_back = request.FILES.get('kyc_back')
        kyc_verified_val = request.data.get('kyc_verified')

        if guest_phone:
            if not (len(guest_phone) == 10 and guest_phone.isdigit()):
                return Response({"detail": "Phone number must be exactly 10 digits."}, status=status.HTTP_400_BAD_REQUEST)
            booking.guest_phone = guest_phone

        if guest_first_name:
            booking.guest_first_name = guest_first_name
        if guest_last_name:
            booking.guest_last_name = guest_last_name
        if guest_email is not None:
            booking.guest_email = guest_email
        if notes is not None:
            booking.notes = notes
        if extra_charges is not None:
            try:
                booking.extra_charges = Decimal(str(extra_charges))
            except (ValueError, TypeError):
                return Response({"detail": "Extra charges must be a number."}, status=status.HTTP_400_BAD_REQUEST)
        if extra_charges_reason is not None:
            booking.extra_charges_reason = extra_charges_reason

        # Get or create customer if it doesn't exist on this booking
        if booking.guest_phone:
            from .models import Customer
            customer = booking.customer
            if not customer:
                if booking.guest_email:
                    customer = Customer.objects.filter(user__username=booking.guest_email).first()
                if not customer:
                    customer = Customer.objects.filter(phone=booking.guest_phone).first()
                
                if not customer:
                    username = booking.guest_email if booking.guest_email else f"{booking.guest_phone}@hotel.com"
                    user_record = User.objects.filter(username=username).first()
                    if not user_record:
                        user_record = User.objects.create_user(
                            username=username,
                            password=username,
                            first_name=booking.guest_first_name,
                            last_name=booking.guest_last_name,
                            email=username
                        )
                    customer = Customer.objects.filter(user=user_record).first()
                    if not customer:
                        customer = Customer.objects.create(
                            user=user_record,
                            phone=booking.guest_phone
                        )
                booking.customer = customer

            if customer:
                if kyc_type:
                    customer.kyc_type = kyc_type
                if kyc_number:
                    customer.kyc_number = kyc_number
                if kyc_front:
                    customer.kyc_front = kyc_front
                if kyc_back:
                    customer.kyc_back = kyc_back
                if kyc_verified_val is not None:
                    new_val = (str(kyc_verified_val).lower() == 'true')
                    if customer.kyc_verified and not new_val:
                        return Response({"detail": "Once verified, verification status cannot be revoked/unverified."}, status=status.HTTP_400_BAD_REQUEST)
                    customer.kyc_verified = new_val
                
                customer.user.first_name = booking.guest_first_name
                customer.user.last_name = booking.guest_last_name
                if booking.guest_email:
                    customer.user.email = booking.guest_email
                customer.user.save()
                customer.save()

        # Handle room reallocation
        room_id = request.data.get('room_id')
        if room_id:
            try:
                new_room = Room.objects.get(id=room_id)
                if new_room.hotel != hotel:
                    return Response({"detail": "Room does not belong to this hotel."}, status=status.HTTP_400_BAD_REQUEST)
                booking.room = new_room
            except Room.DoesNotExist:
                return Response({"detail": "Room not found."}, status=status.HTTP_404_NOT_FOUND)

        # Handle check-in and check-out dates
        cin_date_str = check_in or booking.check_in.strftime('%Y-%m-%d')
        cout_date_str = check_out or booking.check_out.strftime('%Y-%m-%d')

        try:
            check_in_date = datetime.datetime.strptime(cin_date_str, "%Y-%m-%d").date()
            check_out_date = datetime.datetime.strptime(cout_date_str, "%Y-%m-%d").date()
        except ValueError:
            return Response({"detail": "Invalid date format."}, status=status.HTTP_400_BAD_REQUEST)

        if check_in_date >= check_out_date:
            return Response({"detail": "Check-out date must be after check-in date."}, status=status.HTTP_400_BAD_REQUEST)

        overlapping_bookings = Booking.objects.filter(
            room=booking.room,
            check_in__lt=check_out_date,
            check_out__gt=check_in_date
        ).exclude(id=booking.id).exclude(status__in=['Checked_out', 'Cancelled', 'Rejected'])

        if overlapping_bookings.exists():
            return Response({"detail": "This room is already booked for the selected dates."}, status=status.HTTP_400_BAD_REQUEST)

        booking.check_in = check_in_date
        booking.check_out = check_out_date

        if status_val:
            if status_val == 'dirty' or status_val == 'Checked_out':
                booking.status = 'Checked_out'
                booking.checked_out = True
                booking.room.cleanliness = 'dirty'
                booking.room.save()
            else:
                if booking.status == 'Checked_out':
                    # Cleaning room shouldn't undo checkout!
                    booking.room.cleanliness = 'clean'
                    booking.room.save()
                else:
                    booking.status = status_val
                    booking.room.cleanliness = 'clean'
                    booking.room.save()

        if checked_out_val is not None:
            booking.checked_out = checked_out_val
            if checked_out_val:
                booking.status = 'Checked_out'
                booking.room.cleanliness = 'dirty'
                booking.room.save()

        booking.save()

        nights = (booking.check_out - booking.check_in).days
        nights = max(1, nights)
        total_cost = (nights * booking.room.room_type.price) + booking.extra_charges
        
        txns = booking.transactions.all().order_by('created_at')
        txns_data = [{
            "id": t.id,
            "amount": t.amount,
            "payment_method": t.payment_method,
            "receipt_id": t.receipt_id,
            "created_at": timezone.localtime(t.created_at).strftime('%Y-%m-%d %I:%M %p')
        } for t in txns]
        
        total_paid = sum(t.amount for t in txns)
        outstanding_amount = total_cost - total_paid
        actual_advance_status = 'Paid' if total_paid > 0 else 'Unpaid'

        kyc_data = {}
        if booking.customer:
            kyc_data = {
                "kyc_type": booking.customer.kyc_type,
                "kyc_number": booking.customer.kyc_number,
                "kyc_front": booking.customer.kyc_front.url if booking.customer.kyc_front else None,
                "kyc_back": booking.customer.kyc_back.url if booking.customer.kyc_back else None,
                "kyc_verified": booking.customer.kyc_verified,
            }

        return Response({
            "id": booking.id,
            "room_id": booking.room.id,
            "room_number": booking.room.number,
            "room_type": booking.room.room_type.name,
            "guest_first_name": booking.guest_first_name,
            "guest_last_name": booking.guest_last_name,
            "guest_phone": booking.guest_phone,
            "guest_email": booking.guest_email,
            "check_in": booking.check_in.strftime('%Y-%m-%d'),
            "check_in_time": "11:30 AM",
            "check_out": booking.check_out.strftime('%Y-%m-%d'),
            "check_out_time": "11:30 AM",
            "status": booking.status,
            "checked_out": booking.checked_out,
            "rejection_reason": booking.rejection_reason,
            "notes": booking.notes,
            "extra_charges": booking.extra_charges,
            "extra_charges_reason": booking.extra_charges_reason,
            "total_cost": total_cost,
            "advance_paid": total_paid,
            "advance_status": actual_advance_status,
            "outstanding_amount": outstanding_amount,
            "transactions": txns_data,
            "kyc": kyc_data
        }, status=status.HTTP_200_OK)

    elif request.method == 'DELETE':
        reason = request.query_params.get('reason') or request.data.get('reason') or ''
        action = request.query_params.get('action') or request.data.get('action') or ''
        if is_booking_customer:
            if booking.status != 'Booked':
                return Response({"detail": "Cannot cancel a reservation that has already checked in or checked out."}, status=status.HTTP_400_BAD_REQUEST)
            booking.status = 'Cancelled'
            if reason:
                booking.rejection_reason = reason
            booking.save()
            return Response({"detail": "Booking cancelled successfully."}, status=status.HTTP_200_OK)
        else:
            if action == 'cancel':
                booking.status = 'Cancelled'
                default_reason = 'Cancelled by manager'
            else:
                booking.status = 'Rejected'
                default_reason = 'No reason specified'
            
            booking.rejection_reason = reason if reason else default_reason
            
            refund_val = request.query_params.get('refund') or request.data.get('refund')
            if refund_val:
                try:
                    from decimal import Decimal
                    refund_amount = Decimal(str(refund_val))
                    if refund_amount > 0:
                        method = request.query_params.get('method') or request.data.get('method') or 'Cash'
                        from .models import Transaction
                        Transaction.objects.create(
                            booking=booking,
                            amount=-refund_amount,
                            payment_method=method,
                            receipt_id=f"REF-{booking.id}"
                        )
                except Exception:
                    pass
            
            booking.save()
            detail_msg = "Booking cancelled successfully." if booking.status == 'Cancelled' else "Booking rejected successfully."
            return Response({"detail": detail_msg}, status=status.HTTP_200_OK)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_transaction(request, pk):
    """
    Log a new payment transaction associated with a booking.
    """
    user = request.user
    try:
        booking = Booking.objects.get(id=pk)
    except Booking.DoesNotExist:
        return Response({"detail": "Booking not found."}, status=status.HTTP_404_NOT_FOUND)

    hotel = booking.room.hotel
    is_owner = hotel.owner == user
    is_manager = hotel.managers.filter(id=user.id).exists()
    is_booking_customer = booking.customer and hasattr(user, 'customer_profile') and booking.customer == user.customer_profile
    if not (is_owner or is_manager or is_booking_customer):
        return Response({"detail": "Not authorized for this hotel."}, status=status.HTTP_403_FORBIDDEN)

    amount = request.data.get('amount')
    payment_method = request.data.get('payment_method', 'Cash')
    receipt_id = request.data.get('receipt_id')

    if amount is None:
        return Response({"detail": "Amount is required."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        amount_dec = float(amount)
    except ValueError:
        return Response({"detail": "Amount must be a number."}, status=status.HTTP_400_BAD_REQUEST)

    # Create transaction record
    Transaction.objects.create(
        booking=booking,
        amount=amount_dec,
        payment_method=payment_method,
        receipt_id=receipt_id if receipt_id else None
    )

    # Return the updated booking details
    nights = (booking.check_out - booking.check_in).days
    nights = max(1, nights)
    total_cost = (nights * booking.room.room_type.price) + booking.extra_charges
    
    txns = booking.transactions.all().order_by('created_at')
    txns_data = [{
        "id": t.id,
        "amount": t.amount,
        "payment_method": t.payment_method,
        "receipt_id": t.receipt_id,
        "created_at": timezone.localtime(t.created_at).strftime('%Y-%m-%d %I:%M %p')
    } for t in txns]
    
    total_paid = sum(t.amount for t in txns)
    outstanding_amount = total_cost - total_paid
    actual_advance_status = 'Paid' if total_paid > 0 else 'Unpaid'

    kyc_data = {}
    if booking.customer:
        kyc_data = {
            "kyc_type": booking.customer.kyc_type,
            "kyc_number": booking.customer.kyc_number,
            "kyc_front": booking.customer.kyc_front.url if booking.customer.kyc_front else None,
            "kyc_back": booking.customer.kyc_back.url if booking.customer.kyc_back else None,
            "kyc_verified": booking.customer.kyc_verified,
        }

    return Response({
        "id": booking.id,
        "room_id": booking.room.id,
        "room_number": booking.room.number,
        "room_type": booking.room.room_type.name,
        "guest_first_name": booking.guest_first_name,
        "guest_last_name": booking.guest_last_name,
        "guest_phone": booking.guest_phone,
        "guest_email": booking.guest_email,
        "check_in": booking.check_in.strftime('%Y-%m-%d'),
        "check_in_time": "11:30 AM",
        "check_out": booking.check_out.strftime('%Y-%m-%d'),
        "check_out_time": "11:30 AM",
        "status": booking.status,
        "checked_out": booking.checked_out,
        "rejection_reason": booking.rejection_reason,
        "notes": booking.notes,
        "extra_charges": booking.extra_charges,
        "extra_charges_reason": booking.extra_charges_reason,
        "total_cost": total_cost,
        "advance_paid": total_paid,
        "advance_status": actual_advance_status,
        "outstanding_amount": outstanding_amount,
        "transactions": txns_data,
        "kyc": kyc_data
    }, status=status.HTTP_201_CREATED)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_room_cleanliness(request, pk):
    """
    Update a room's cleanliness status directly (independent of bookings).
    """
    user = request.user
    try:
        room = Room.objects.get(id=pk)
    except Room.DoesNotExist:
        return Response({"detail": "Room not found."}, status=status.HTTP_404_NOT_FOUND)

    hotel = room.hotel
    is_owner = hotel.owner == user
    is_manager = hotel.managers.filter(id=user.id).exists()
    if not (is_owner or is_manager):
        return Response({"detail": "Not authorized for this hotel."}, status=status.HTTP_403_FORBIDDEN)

    cleanliness = request.data.get('cleanliness')
    if cleanliness not in ['clean', 'dirty']:
        return Response({"detail": "Invalid cleanliness value. Choose 'clean' or 'dirty'."}, status=status.HTTP_400_BAD_REQUEST)

    room.cleanliness = cleanliness
    room.save()

    return Response({
        "id": room.id,
        "number": room.number,
        "cleanliness": room.cleanliness
    }, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_transactions_excel(request):
    """
    Generate and download a formatted Excel spreadsheet (.xlsx) of transactions
    filtered by today, yesterday, or custom date range in India timezone (+05:30).
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    import datetime

    user = request.user
    hotel_code = request.query_params.get('hotel_code')
    if not hotel_code:
        return Response({"detail": "hotel_code query parameter is required."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        hotel = Hotel.objects.get(code=hotel_code)
    except Hotel.DoesNotExist:
        return Response({"detail": "Hotel not found."}, status=status.HTTP_404_NOT_FOUND)

    # Authorization Check
    is_owner = hotel.owner == user
    is_manager = hotel.managers.filter(id=user.id).exists()
    if not (is_owner or is_manager):
        return Response({"detail": "Not authorized for this hotel."}, status=status.HTTP_403_FORBIDDEN)

    # Fetch bookings for this hotel
    bookings = Booking.objects.filter(room__hotel=hotel)
    
    # We want to pull all transactions for these bookings
    txns = Transaction.objects.filter(booking__in=bookings).order_by('created_at')

    # Get local dates (Indian Standard Time UTC + 5:30)
    tz_shift = datetime.timedelta(hours=5, minutes=30)
    
    # Parse filter params
    txn_filter = request.query_params.get('filter', 'all')
    start_date_str = request.query_params.get('start_date')
    end_date_str = request.query_params.get('end_date')

    today_local = (timezone.now() + tz_shift).date()
    yesterday_local = today_local - datetime.timedelta(days=1)

    filtered_txns = []
    for t in txns:
        local_created = t.created_at + tz_shift
        local_date = local_created.date()
        
        # Apply filter matching frontend logic
        if txn_filter == 'today':
            if local_date != today_local:
                continue
        elif txn_filter == 'yesterday':
            if local_date != yesterday_local:
                continue
        elif txn_filter == 'custom':
            if start_date_str:
                try:
                    start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
                    if local_date < start_date:
                        continue
                except ValueError:
                    pass
            if end_date_str:
                try:
                    end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()
                    if local_date > end_date:
                        continue
                except ValueError:
                    pass
        
        b = t.booking
        filtered_txns.append({
            "txn": t,
            "local_datetime": local_created,
            "guest_name": f"{b.guest_first_name} {b.guest_last_name}",
            "room_info": f"Room {b.room.number} ({b.room.room_type.name})",
            "incidental_reason": b.extra_charges_reason,
            "amount": float(t.amount)
        })

    # Summary calculations
    unique_bookings = set(item['txn'].booking_id for item in filtered_txns if item['amount'] > 0)
    unique_refunds = set(item['txn'].booking_id for item in filtered_txns if item['amount'] < 0)
    
    total_booking_amt = sum(item['amount'] for item in filtered_txns if item['amount'] > 0)
    total_refund_amt = sum(abs(item['amount']) for item in filtered_txns if item['amount'] < 0)
    net_income = total_booking_amt - total_refund_amt

    # Create Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions Report"
    
    # Ensure gridlines are visible
    ws.views.sheetView[0].showGridLines = True

    # Styling Palettes
    navy_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    title_font = Font(name="Calibri", size=16, bold=True, color="1E1B4B")
    subtitle_font = Font(name="Calibri", size=11, italic=True, color="475569")
    bold_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=11)
    
    green_text = Font(name="Calibri", size=11, bold=True, color="16A34A")
    red_text = Font(name="Calibri", size=11, bold=True, color="DC2626")
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    double_bottom_border = Border(
        top=Side(style='thin', color='94A3B8'),
        bottom=Side(style='double', color='1E293B')
    )

    # 1. Title Block
    ws['A1'] = "FINANCIAL TRANSACTIONS REPORT"
    ws['A1'].font = title_font
    ws['A1'].alignment = left_align
    
    ws['A2'] = f"Hotel Name: {hotel.name} ({hotel.code})"
    ws['A2'].font = subtitle_font
    
    range_desc = "All Recorded Transactions"
    if txn_filter == 'today':
        range_desc = f"Today ({today_local.strftime('%Y-%m-%d')})"
    elif txn_filter == 'yesterday':
        range_desc = f"Yesterday ({yesterday_local.strftime('%Y-%m-%d')})"
    elif txn_filter == 'custom':
        range_desc = f"Range: {start_date_str or 'Start'} to {end_date_str or 'End'}"
    ws['A3'] = f"Report Interval: {range_desc}"
    ws['A3'].font = subtitle_font
    
    ws['A4'] = f"Generated On: {(timezone.now() + tz_shift).strftime('%Y-%m-%d %I:%M %p')}"
    ws['A4'].font = subtitle_font

    # 2. Summary Block
    ws['A6'] = "FINANCIAL SUMMARY"
    ws['A6'].font = Font(name="Calibri", size=12, bold=True, color="1E293B")
    
    summary_headers = ["Metric", "Count", "Total Value (INR)"]
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = ws.cell(row=7, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = navy_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    metrics = [
        ("Booking Payments", len(unique_bookings), total_booking_amt, green_text),
        ("Refunds Issued", len(unique_refunds), -total_refund_amt, red_text),
        ("Net Settlements", "", net_income, bold_font)
    ]
    
    for row_offset, (label, count, val, val_style) in enumerate(metrics, start=8):
        c_label = ws.cell(row=row_offset, column=1, value=label)
        c_count = ws.cell(row=row_offset, column=2, value=count)
        c_val = ws.cell(row=row_offset, column=3, value=val)
        
        c_label.font = bold_font
        c_count.font = normal_font
        c_val.font = val_style if val_style else normal_font
        
        c_label.alignment = left_align
        c_count.alignment = center_align
        c_val.alignment = right_align
        
        c_label.border = thin_border
        c_count.border = thin_border
        c_val.border = thin_border
        
        c_val.number_format = '[$₹-437] #,##0.00'

    # 3. Detailed Transactions Table
    start_row = 13
    ws.cell(row=start_row-1, column=1, value="DETAILED TRANSACTION LEDGER").font = Font(name="Calibri", size=12, bold=True, color="1E293B")
    
    headers = [
        "Date & Time", 
        "Guest Name", 
        "Room Info", 
        "Type", 
        "Method", 
        "Receipt/Txn ID", 
        "Amount (INR)"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = navy_fill
        cell.alignment = center_align
        cell.border = thin_border

    current_row = start_row + 1
    for idx, item in enumerate(filtered_txns):
        t = item['txn']
        is_refund = item['amount'] < 0
        type_str = "Refund" if is_refund else "Booking Payment"
        
        r_dt = ws.cell(row=current_row, column=1, value=item['local_datetime'].strftime('%Y-%m-%d %I:%M %p'))
        r_name = ws.cell(row=current_row, column=2, value=item['guest_name'])
        
        room_desc = item['room_info']
        if item['incidental_reason']:
            room_desc += f" [Incid: {item['incidental_reason']}]"
        r_room = ws.cell(row=current_row, column=3, value=room_desc)
        
        r_type = ws.cell(row=current_row, column=4, value=type_str)
        r_method = ws.cell(row=current_row, column=5, value=t.payment_method)
        r_receipt = ws.cell(row=current_row, column=6, value=t.receipt_id or "-")
        r_amt = ws.cell(row=current_row, column=7, value=item['amount'])
        
        r_dt.alignment = center_align
        r_name.alignment = left_align
        r_room.alignment = left_align
        r_type.alignment = center_align
        r_method.alignment = center_align
        r_receipt.alignment = center_align
        r_amt.alignment = right_align
        
        for c in range(1, 8):
            cell = ws.cell(row=current_row, column=c)
            cell.border = thin_border
            cell.font = normal_font
            
        if is_refund:
            r_type.font = Font(name="Calibri", size=11, bold=True, color="DC2626")
            r_amt.font = Font(name="Calibri", size=11, color="DC2626")
        else:
            r_type.font = normal_font
            r_amt.font = Font(name="Calibri", size=11, color="16A34A")
            
        r_amt.number_format = '[$₹-437] #,##0.00'
        current_row += 1

    if filtered_txns:
        ws.cell(row=current_row, column=1, value="Total Net Income").font = bold_font
        ws.cell(row=current_row, column=1).alignment = left_align
        
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        
        for c in range(1, 7):
            cell = ws.cell(row=current_row, column=c)
            cell.border = double_bottom_border
            
        r_total_val = ws.cell(row=current_row, column=7, value=f"=SUM(G{start_row+1}:G{current_row-1})")
        r_total_val.font = bold_font
        r_total_val.alignment = right_align
        r_total_val.border = double_bottom_border
        r_total_val.number_format = '[$₹-437] #,##0.00'
    else:
        ws.cell(row=current_row, column=1, value="No transactions found for the selected range.").font = subtitle_font
        ws.cell(row=current_row, column=1).alignment = center_align
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        for c in range(1, 8):
            ws.cell(row=current_row, column=c).border = thin_border

    # Auto-adjust column dimensions
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # check if cell is merged
            is_merged = False
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    # if it's not the top-left cell, ignore it for width
                    if cell.coordinate != merged_range.start_cell.coordinate:
                        is_merged = True
                    break
            if is_merged:
                continue
                
            val_str = str(cell.value or '')
            if '₹' in val_str or '=' in val_str:
                max_len = max(max_len, 14)
            else:
                max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Build response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"transactions_{hotel_code}_{txn_filter}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

